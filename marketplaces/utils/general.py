from openpyxl import load_workbook
import requests
from io import BytesIO
import logging
import argparse
from datetime import datetime
from utils.logs import Logs

class GeneralUtils:
    """
    A utility class for general data processing tasks such as reading Excel files from URLs
    and SQL files from the filesystem, with integrated structured logging.
    """
    def __init__(self, logger_name: str, log_level=logging.INFO):
        """
        Initializes the GeneralUtils class with a specific logger name and log level,
        setting up structured logging for all its methods.
        """
        self.logging_instance = Logs(name=logger_name, log_level=log_level)
        self.logger = self.logging_instance.get_logger()

    def read_excel_from_url(self, url: str) -> dict:
        """
        Downloads an Excel file (.xlsx format only) from the given URL and reads its content into a list of lists.
        Not pd df due to memory savings.

        Args:
        url (str): The URL of the .xlsx Excel file to be downloaded and read.

        Returns:
        list: A list of lists, where each inner list represents a row in the Excel sheet, if successful.

        Note:
        It works only with .xlsx files and not with older .xls formats.
        """
        self.logger.info(event="reading_excel_from_url_started", url=url)
        try:
            response = requests.get(url)
            if response.status_code == 200:
                workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
                all_sheets_data = {}
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
                    all_sheets_data[sheet_name] = sheet_data
                self.logger.info(event="reading_excel_from_url_successful", url=url)
                return all_sheets_data
            else:
                self.logger.error(event="excel_file_retrieval_failed", http_status=response.status_code, http_reason=response.reason)
                raise
        except requests.RequestException as e:
            self.logger.error(event="http_request_error", error=str(e))
            raise
        except Exception as e:
            self.logger.error(event="unexpected_error_reading_excel", error=str(e))
            raise

    def read_sql_file(self, filepath: str, **params) -> str:
        """
        Reads a SQL file from the local file system and returns its content as a string along with a success flag.

        Args:
        filepath (str): The file path of the SQL file to be read.
        **params: Arbitrary keyword arguments representing parameters to format the SQL query.

        Returns:
        str: A str with sql query text.
        """
        self.logger.info(event="reading_sql_file_started", filepath=filepath)
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                self.logger.info(event="reading_sql_file_successful", filepath=filepath)
                query = file.read().format(**params)
                if query:
                    return query
                else:
                    self.logger.error(event="sql_file_is_empty", filepath=filepath)
                    exit()
        except FileNotFoundError:
            self.logger.error(event="sql_file_not_found", filepath=filepath)
            raise
        except Exception as e:
            self.logger.error(event="unexpected_error_reading_sql_file", error=str(e))
            raise